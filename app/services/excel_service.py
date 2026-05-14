from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_excel(results: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcription Results"

    # Purple theme styling
    header_fill = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D4B8E0"),
        right=Side(style="thin", color="D4B8E0"),
        top=Side(style="thin", color="D4B8E0"),
        bottom=Side(style="thin", color="D4B8E0"),
    )
    alt_fill = PatternFill(start_color="F5F0F7", end_color="F5F0F7", fill_type="solid")

    # Headers
    headers = ["#", "Filename", "Transcript", "LLM Analysis"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for idx, item in enumerate(results, 1):
        row = idx + 1
        cells = [
            ws.cell(row=row, column=1, value=idx),
            ws.cell(row=row, column=2, value=item.get("filename", "")),
            ws.cell(row=row, column=3, value=item.get("transcript", "")),
            ws.cell(row=row, column=4, value=item.get("analysis", "")),
        ]
        for cell in cells:
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
